# 엑셀 다루기 
 # openpyxl
 # pip install openpyxl
 import openpyxl
 # https://openpyxl.readthedocs.io/
 # import openpyxl
  
 wb = openpyxl.load_workbook("score.xlsx")
 ws = wb.active
 # wb = openpyxl.load_workbook("score.xlsx")
 # ws = wb.active
  
 for r in ws.rows:
 	row_index = r[0].row
 	kor = r[1].value
 	eng = r[1].value
	math = r[1].value
	sum = kor+eng+math
# for r in ws.rows:
# 	row_index = r[0].row
# 	kor = r[1].value
# 	eng = r[2].value
# 	math = r[3].value
# 	sum = kor+eng+math
  
	ws.cell(row=row_index, column=5).value = sum
# 	ws.cell(row=row_index, column=5).value = sum
  
	print(kor,eng,math,sum)

wb.save("score2.xlsx")
wb.close
# wb.save("score2.xlsx")
# wb.close
  